import pandas as pd
import os
from datetime import datetime
from models import Attendance, Student, Course

def export_attendance_to_excel(course_code=None, filename=None):
    """
    Export attendance records to Excel
    If course_code is provided, export only that course
    Otherwise, export all attendance records
    """
    if filename is None:
        today = datetime.now().strftime('%Y-%m-%d')
        if course_code:
            filename = os.path.join("exports", f"attendance_{course_code}_{today}.xlsx")
        else:
            filename = os.path.join("exports", f"attendance_all_{today}.xlsx")
    
    # Get attendance records
    if course_code:
        records = Attendance.get_attendance_by_course(course_code)
        course = Course.get_course_by_code(course_code)
        sheet_name = f"{course_code}"
    else:
        records = Attendance.get_all_attendance()
        sheet_name = "All Attendance"
    
    # Convert to list of dictionaries
    data = []
    for record in records:
        data.append({
            'Student ID': record['student_id'],
            'Student Name': record['name'] if 'name' in record.keys() else '',
            'Course Code': record['course_code'],
            'Course Name': record['course_name'] if 'course_name' in record.keys() else '',
            'Date': record['date'],
            'Time': record['time'],
            'Status': record['status'],
            'Marked At': record['marked_at']
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    if df.empty:
        print("No attendance records to export")
        return None
    
    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get the worksheet
        worksheet = writer.sheets[sheet_name]
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    print(f"Attendance exported to {filename}")
    return filename

def export_student_attendance_summary(filename=None):
    """
    Export student attendance summary with percentages
    """
    if filename is None:
        today = datetime.now().strftime('%Y-%m-%d')
        filename = os.path.join("exports", f"attendance_summary_{today}.xlsx")
    
    students = Student.get_all_students()
    courses = Course.get_all_courses()
    
    data = []
    
    for student in students:
        student_data = {
            'Student ID': student['student_id'],
            'Student Name': student['name'],
            'Email': student['email']
        }
        
        for course in courses:
            percentage = Attendance.get_student_attendance_percentage(
                student['student_id'], 
                course['course_code']
            )
            student_data[f"{course['course_code']} (%)"] = percentage
        
        data.append(student_data)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    if df.empty:
        print("No student data to export")
        return None
    
    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Attendance Summary', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Attendance Summary']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Highlight low attendance (below threshold)
        from openpyxl.styles import PatternFill
        from models import Settings
        
        threshold = Settings.get_min_attendance_percentage()
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
            for col in range(4, len(df.columns) + 1):  # Start from column 4 (after basic info)
                cell = worksheet.cell(row=row, column=col)
                try:
                    if float(cell.value) < threshold:
                        cell.fill = red_fill
                except:
                    pass
    
    print(f"Attendance summary exported to {filename}")
    return filename
